import yfinance as yf
import openpyxl
from openpyxl.utils import get_column_letter

# Lista de tickers
tickers = ["MXRF11.SA", "BBAS3.SA", "BBSE3.SA", "CMIG4.SA", "KLBN11.SA", "PETR4.SA", "PFRM3.SA", "PGMN3.SA", 
           "VALE3.SA", "BTLG11.SA", "XPLG11.SA", "XPSF11.SA", "VGHF11.SA", "CPTS11.SA", "VILG11.SA", "KNSC11.SA"]

# Caminho da planilha (ajuste para seu usuário)
caminho_arquivo = "/home/eduardo/Área de trabalho/Financeiro/dados.xlsx"

# Função para calcular o valor intrínseco e margem de segurança
def calcular_margem_seguranca(ticker, taxa_crescimento=0.05):
    try:
        acao = yf.Ticker(ticker)
        info = acao.info

        preco_atual = info.get('regularMarketPrice', 0)
        eps = info.get('epsTrailingTwelveMonths', 0)

        # Tentando pegar a taxa de crescimento projetada (se disponível)
        taxa_crescimento = info.get('earningsQuarterlyGrowth', None)  # A chave 'growth' pode não estar presente

        # Se a taxa de crescimento não estiver disponível, usa 5% como padrão
        if taxa_crescimento is None:
            print(f"Taxa de crescimento projetada não disponível para {ticker}. Usando 5% como padrão.")
            taxa_crescimento = 0.05

        if eps == 0:
            return preco_atual, eps, 0, 0  # Evita erro se não houver EPS disponível

        valor_intrinseco = eps * (8.5 + 2 * taxa_crescimento)
        margem_seguranca = valor_intrinseco - preco_atual

        return preco_atual, eps, valor_intrinseco, margem_seguranca

    except Exception as e:
        print(f"Erro ao buscar {ticker}: {e}")
        return 0, 0, 0, 0

# Função para atualizar a planilha sem duplicar os tickers
def atualizar_planilha():
    try:
        wb = openpyxl.load_workbook(caminho_arquivo)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Ticker", "Preço Atual", "EPS", "Valor Intrínseco", "Margem de Segurança"])

    # Obtém os tickers existentes na planilha para evitar duplicação
    tickers_existentes = {sheet.cell(row=i, column=1).value: i for i in range(2, sheet.max_row + 1)}

    # Atualiza ou adiciona os dados dos tickers
    for ticker in tickers:
        preco_atual, eps, valor_intrinseco, margem_seguranca = calcular_margem_seguranca(ticker)

        if ticker in tickers_existentes:
            row = tickers_existentes[ticker]
            sheet.cell(row=row, column=2, value=preco_atual)
            sheet.cell(row=row, column=3, value=eps)
            sheet.cell(row=row, column=4, value=valor_intrinseco)
            sheet.cell(row=row, column=5, value=margem_seguranca)
        else:
            sheet.append([ticker, preco_atual, eps, valor_intrinseco, margem_seguranca])

    # Salva a planilha com as atualizações
    wb.save(caminho_arquivo)
    print(f"Dados atualizados em {caminho_arquivo}")

# Executa a atualização
atualizar_planilha()
